"""
Voice Library Import Script

Import MiniMax voice library data from Excel file into database.

Usage:
    cd main/live-agent-api
    
    # Using environment variable
    export DATABASE_URL="postgresql+asyncpg://user:pass@localhost:5432/live_agent"
    uv run python scripts/import_voice_library.py --dry-run
    
    # Or using command line argument
    uv run python scripts/import_voice_library.py --db "postgresql+asyncpg://user:pass@localhost:5432/live_agent" --dry-run

Requirements:
    - openpyxl (for Excel reading)
    - Add to pyproject.toml dev dependencies: "openpyxl>=3.1.0"
"""

import asyncio
import os
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from dotenv import load_dotenv

# Load .env file from project root
env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

from datetime import datetime, timezone
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker, DeclarativeBase, Mapped, mapped_column
from sqlalchemy import select, String, Text, TIMESTAMP, text
from sqlalchemy.dialects.postgresql import JSONB

from ulid import ULID
from loguru import logger


# System user ID for voice library voices
SYSTEM_VOICE_LIBRARY_OWNER = "system_voice_library"


async def ensure_system_user(session: AsyncSession):
    """Ensure system_voice_library user exists in user table"""
    # Check if user exists
    result = await session.execute(
        text("SELECT 1 FROM \"user\" WHERE user_id = :user_id"),
        {"user_id": SYSTEM_VOICE_LIBRARY_OWNER}
    )
    if result.scalar_one_or_none() is None:
        # Create system user
        await session.execute(
            text("""
                INSERT INTO "user" (user_id, username, password)
                VALUES (:user_id, :username, :password)
                ON CONFLICT (user_id) DO NOTHING
            """),
            {
                "user_id": SYSTEM_VOICE_LIBRARY_OWNER,
                "username": SYSTEM_VOICE_LIBRARY_OWNER,
                "password": "not_for_login"
            }
        )
        await session.commit()
        logger.info(f"Created system user: {SYSTEM_VOICE_LIBRARY_OWNER}")


# ==================== Standalone ORM Model (avoid importing app config) ====================

class Base(DeclarativeBase):
    pass


def utc_now():
    return datetime.now(timezone.utc)


class VoiceModel(Base):
    """Standalone VoiceModel for import script"""
    __tablename__ = "voices"

    id: Mapped[int] = mapped_column(primary_key=True, autoincrement=True)
    voice_id: Mapped[str] = mapped_column(String(50), unique=True, nullable=False)
    reference_id: Mapped[str | None] = mapped_column(String(100), nullable=True)
    owner_id: Mapped[str] = mapped_column(String(50), nullable=False)
    category: Mapped[str] = mapped_column(String(20), nullable=False, default="clone")
    provider: Mapped[str] = mapped_column(String(20), nullable=False, default="fishspeech")
    name: Mapped[str] = mapped_column(String(100), nullable=False)
    desc: Mapped[str] = mapped_column(Text, nullable=False)
    tags: Mapped[dict] = mapped_column(JSONB, nullable=False, default=dict)
    sample_url: Mapped[str | None] = mapped_column(Text, nullable=True)
    sample_text: Mapped[str | None] = mapped_column(Text, nullable=True)
    created_at: Mapped[datetime] = mapped_column(TIMESTAMP(timezone=True), default=utc_now, nullable=False)
    updated_at: Mapped[datetime] = mapped_column(TIMESTAMP(timezone=True), default=utc_now, nullable=False)


# Excel column mapping (0-indexed)
COL_LANGUAGE = 0      # A: 语种名
COL_NAME = 1          # B: Name
COL_REFERENCE_ID = 2  # C: voice_id (MiniMax's voice_id)
COL_GENDER = 3        # D: Gender标签
COL_AGE = 4           # E: Age标签
COL_STYLE = 5         # F: Style标签
COL_ACCENT = 6        # G: Accent标签
COL_DESCRIPTION = 7   # H: Description


# Language code mapping
LANGUAGE_MAP = {
    "english": "en",
    "chinese": "zh",
    "japanese": "ja",
    "korean": "ko",
    "spanish": "es",
    "french": "fr",
    "german": "de",
    "italian": "it",
    "portuguese": "pt",
    "russian": "ru",
    "arabic": "ar",
    "hindi": "hi",
    "thai": "th",
    "vietnamese": "vi",
    "indonesian": "id",
    "malay": "ms",
    "turkish": "tr",
    "dutch": "nl",
    "polish": "pl",
}

# Gender mapping
GENDER_MAP = {
    "male": "male",
    "female": "female",
}

# Age mapping (handles various formats from Excel)
AGE_MAP = {
    "youth": "youth",
    "young adult": "young_adult",
    "young_adult": "young_adult",
    "youngadult": "young_adult",
    "adult": "adult",
    "middle aged": "middle_aged",
    "middle-aged": "middle_aged",
    "middle_aged": "middle_aged",
    "middleaged": "middle_aged",
    "middle age": "middle_aged",
    "senior": "senior",
}


def generate_voice_id() -> str:
    """Generate unique voice ID with 'voice_' prefix"""
    return f"voice_{ULID()}"


def normalize_gender(value: str) -> str | None:
    """Normalize gender value"""
    if not value:
        return None
    return GENDER_MAP.get(value.lower().strip())


def normalize_age(value: str) -> str | None:
    """Normalize age value"""
    if not value:
        return None
    return AGE_MAP.get(value.lower().strip())


def normalize_language(value: str) -> str | None:
    """Normalize language to ISO code"""
    if not value:
        return None
    return LANGUAGE_MAP.get(value.lower().strip(), value.lower().strip())


def generate_sample_url(voice_id: str, s3_base_url: str, bucket_name: str) -> str:
    """
    Generate placeholder sample URL following clone voice format.
    
    Format: {S3_PUBLIC_BASE_URL}/{S3_BUCKET_NAME}/voice_samples/{voice_id}.wav
    Example: http://18.143.177.88:19000/live-agent/voice_samples/voice_xxx.wav
    """
    return f"{s3_base_url}/{bucket_name}/voice_samples/{voice_id}.wav"


def parse_excel_row(row: tuple, s3_base_url: str, bucket_name: str) -> dict | None:
    """
    Parse a single Excel row into voice data dict.
    
    Returns None if row is header or invalid.
    """
    # Skip header row (check if first cell is "语种名" or similar header)
    first_cell = str(row[COL_LANGUAGE] or "").strip()
    if not first_cell or first_cell.lower() in ["语种名", "language", "语种"]:
        return None
    
    # Extract values
    language_name = str(row[COL_LANGUAGE] or "").strip()
    name = str(row[COL_NAME] or "").strip()
    reference_id = str(row[COL_REFERENCE_ID] or "").strip()
    gender = str(row[COL_GENDER] or "").strip()
    age = str(row[COL_AGE] or "").strip()
    style = str(row[COL_STYLE] or "").strip()
    accent = str(row[COL_ACCENT] or "").strip()
    description = str(row[COL_DESCRIPTION] or "").strip()
    
    # Skip if essential fields are missing
    if not name or not reference_id:
        logger.warning(f"Skipping row with missing name or reference_id: {row}")
        return None
    
    # Generate our voice_id
    voice_id = generate_voice_id()
    
    # Build tags dict
    tags = {}
    
    normalized_gender = normalize_gender(gender)
    if normalized_gender:
        tags["gender"] = normalized_gender
    
    normalized_age = normalize_age(age)
    if normalized_age:
        tags["age"] = normalized_age
    
    normalized_language = normalize_language(language_name)
    if normalized_language:
        tags["language"] = normalized_language
    
    if style:
        tags["style"] = style.lower().replace(" ", "_")
    
    if accent:
        tags["accent"] = accent
    
    if description:
        tags["description"] = description
    
    return {
        "voice_id": voice_id,
        "reference_id": reference_id,
        "owner_id": SYSTEM_VOICE_LIBRARY_OWNER,
        "name": name,
        "desc": description or f"{name} - {style}" if style else name,
        "category": "library",
        "provider": "minimax",
        "tags": tags,
        "sample_url": generate_sample_url(voice_id, s3_base_url, bucket_name),
        "sample_text": None,
    }


async def import_voices(
    excel_path: str, 
    database_url: str,
    s3_base_url: str,
    bucket_name: str,
    dry_run: bool = False
):
    """
    Import voices from Excel file to database.
    
    Args:
        excel_path: Path to Excel file
        database_url: Database connection URL
        s3_base_url: S3 public base URL for sample files
        bucket_name: S3 bucket name
        dry_run: If True, only print what would be imported without actually importing
    """
    logger.info(f"Loading Excel file: {excel_path}")
    
    # Load Excel workbook
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    # Parse all rows
    voices_to_import = []
    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        voice_data = parse_excel_row(row, s3_base_url, bucket_name)
        if voice_data:
            voices_to_import.append(voice_data)
            logger.debug(f"Parsed row {row_num}: {voice_data['name']}")
    
    logger.info(f"Found {len(voices_to_import)} voices to import")
    
    if dry_run:
        logger.info("=== DRY RUN MODE ===")
        for i, voice in enumerate(voices_to_import, 1):
            print(f"\n{i}. {voice['name']}")
            print(f"   voice_id: {voice['voice_id']}")
            print(f"   reference_id: {voice['reference_id']}")
            print(f"   provider: {voice['provider']}")
            print(f"   tags: {voice['tags']}")
            print(f"   sample_url: {voice['sample_url']}")
        logger.info(f"Would import {len(voices_to_import)} voices")
        return
    
    # Create database connection
    engine = create_async_engine(database_url, echo=False)
    async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    
    # Ensure system user exists first
    async with async_session() as session:
        await ensure_system_user(session)
    
    # Import voices one by one with individual transactions (upsert)
    inserted_count = 0
    updated_count = 0
    error_count = 0
    
    for voice_data in voices_to_import:
        async with async_session() as session:
            try:
                # Check if voice with same reference_id already exists
                result = await session.execute(
                    select(VoiceModel).where(
                        VoiceModel.reference_id == voice_data["reference_id"],
                        VoiceModel.provider == voice_data["provider"]
                    )
                )
                existing_voice = result.scalar_one_or_none()
                
                if existing_voice:
                    # Update existing voice (upsert)
                    existing_voice.name = voice_data["name"]
                    existing_voice.desc = voice_data["desc"]
                    existing_voice.tags = voice_data["tags"]
                    existing_voice.sample_url = voice_data["sample_url"]
                    existing_voice.category = voice_data["category"]
                    await session.commit()
                    updated_count += 1
                    logger.info(f"Updated: {voice_data['name']} ({existing_voice.voice_id})")
                else:
                    # Create new voice
                    voice = VoiceModel(**voice_data)
                    session.add(voice)
                    await session.commit()
                    inserted_count += 1
                    logger.info(f"Added: {voice_data['name']} ({voice_data['voice_id']})")
                
            except Exception as e:
                await session.rollback()
                logger.error(f"Failed to import {voice_data['name']}: {e}")
                error_count += 1
    
    await engine.dispose()
    
    logger.info(f"Import complete: {inserted_count} inserted, {updated_count} updated, {error_count} errors")


async def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Import MiniMax voice library from Excel")
    parser.add_argument(
        "excel_file",
        nargs="?",
        default="migrations/MM voice library.xlsx",
        help="Path to Excel file (default: migrations/MM voice library.xlsx)"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be imported without actually importing"
    )
    parser.add_argument(
        "--db",
        dest="database_url",
        default=None,
        help="Database URL (default: from .env DATABASE_URL)"
    )
    parser.add_argument(
        "--s3-base-url",
        dest="s3_base_url",
        default=None,
        help="S3 public base URL (default: from .env S3_PUBLIC_BASE_URL)"
    )
    parser.add_argument(
        "--bucket",
        dest="bucket_name",
        default=None,
        help="S3 bucket name (default: from .env S3_BUCKET_NAME)"
    )
    
    args = parser.parse_args()
    
    # Use args if provided, otherwise fallback to env vars loaded from .env
    database_url = args.database_url or os.getenv("DATABASE_URL")
    s3_base_url = args.s3_base_url or os.getenv("S3_PUBLIC_BASE_URL")
    bucket_name = args.bucket_name or os.getenv("S3_BUCKET_NAME")
    
    # Log loaded config
    logger.info(f"Config loaded from .env:")
    logger.info(f"  DATABASE_URL: {database_url[:50]}..." if database_url and len(database_url) > 50 else f"  DATABASE_URL: {database_url}")
    logger.info(f"  S3_PUBLIC_BASE_URL: {s3_base_url}")
    logger.info(f"  S3_BUCKET_NAME: {bucket_name}")
    
    # Validate required config
    if not s3_base_url:
        logger.error("S3_PUBLIC_BASE_URL is required. Set it in .env or via --s3-base-url argument")
        sys.exit(1)
    if not bucket_name:
        logger.error("S3_BUCKET_NAME is required. Set it in .env or via --bucket argument")
        sys.exit(1)
    
    # Database URL only required for non-dry-run mode
    if not args.dry_run and not database_url:
        logger.error("DATABASE_URL is required. Set it in .env or via --db argument")
        sys.exit(1)
    
    # Resolve path
    excel_path = Path(args.excel_file)
    if not excel_path.is_absolute():
        excel_path = Path(__file__).parent.parent / excel_path
    
    if not excel_path.exists():
        logger.error(f"Excel file not found: {excel_path}")
        sys.exit(1)
    
    await import_voices(
        str(excel_path), 
        database_url or "",
        s3_base_url or "",
        bucket_name or "",
        dry_run=args.dry_run
    )


if __name__ == "__main__":
    asyncio.run(main())
